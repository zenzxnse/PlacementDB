#!/usr/bin/env python3

"""Validate the PlacementDB source workbook without emitting record content."""

from __future__ import annotations

import argparse
import hashlib
import json
import sys
from pathlib import Path
from zipfile import BadZipFile, ZipFile

MAX_FILE_BYTES = 20 * 1024 * 1024
MAX_ARCHIVE_ENTRIES = 512
MAX_UNCOMPRESSED_BYTES = 100 * 1024 * 1024

EXPECTED_TABLES = {
    "QuestionBankTable": (
        "Question ID",
        "Experience ID",
        "Placement year",
        "Year basis",
        "SRM affiliation",
        "Company",
        "Role",
        "Campus / scope",
        "Round",
        "Question type",
        "Topic",
        "Remembered question / prompt",
        "Wording fidelity",
        "Source confidence",
        "Source ID",
        "Source URL",
        "Notes",
    ),
    "PracticeBankTable": (
        "Practice ID",
        "Year / reference",
        "Company",
        "Stage",
        "Topic",
        "Practice prompt / preparation item",
        "Provenance class",
        "Source ID",
        "Source confidence",
        "Source URL",
        "Notes",
    ),
    "ExperiencesTable": (
        "Experience ID",
        "Placement year",
        "Year basis",
        "Affiliation status",
        "Company",
        "Role",
        "Campus / scope",
        "Channel",
        "Rounds",
        "Selection funnel",
        "Outcome",
        "Source ID",
        "Source URL",
        "Source confidence",
    ),
    "PlacementStatsTable": (
        "Record ID",
        "Academic year",
        "Scope",
        "Program / metric",
        "Graduating / strength",
        "Placed",
        "Placement rate",
        "Median salary INR",
        "Higher studies",
        "Offers / listed records",
        "Recruiters / companies",
        "Highest CTC LPA",
        "Average CTC LPA",
        "Entrepreneurs",
        "Source ID",
        "Record status",
        "Notes",
    ),
    "RecruiterCountsTable": (
        "Academic year / batch",
        "Department",
        "Sequence",
        "Process date",
        "Recruiter",
        "Offers / placements",
        "Program scope",
        "Source ID",
    ),
    "SourcesTable": (
        "Source ID",
        "Title",
        "Source type",
        "Publisher / owner",
        "Published / event date",
        "Reliability",
        "Coverage",
        "URL",
        "Caveat / extraction note",
    ),
    "LeadsTable": (
        "Lead ID",
        "Lead",
        "Blocker",
        "Why it matters / next action",
        "Source ID(s)",
        "Priority",
    ),
}

FORBIDDEN_ARCHIVE_PARTS = (
    "vbaproject",
    "xl/embeddings/",
    "xl/externallinks/",
    "activex",
    "customui/",
)


def ParseArguments() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Validate a PlacementDB XLSX workbook without loading it."
    )
    parser.add_argument("workbook", type=Path)
    parser.add_argument("--report", type=Path)
    return parser.parse_args()


def PreflightArchive(path: Path) -> list[str]:
    errors: list[str] = []
    if path.suffix.lower() != ".xlsx":
        errors.append("workbook must use the .xlsx format")
    if not path.is_file():
        return errors + ["workbook does not exist or is not a regular file"]
    if path.stat().st_size > MAX_FILE_BYTES:
        errors.append("workbook exceeds the 20 MiB compressed size limit")
    try:
        with ZipFile(path) as archive:
            entries = archive.infolist()
            if len(entries) > MAX_ARCHIVE_ENTRIES:
                errors.append("workbook contains too many archive entries")
            if sum(entry.file_size for entry in entries) > MAX_UNCOMPRESSED_BYTES:
                errors.append("workbook exceeds the 100 MiB expanded size limit")
            names = [entry.filename.lower() for entry in entries]
            for marker in FORBIDDEN_ARCHIVE_PARTS:
                if any(marker in name for name in names):
                    errors.append(f"workbook contains forbidden active part: {marker}")
    except BadZipFile:
        errors.append("workbook is not a valid ZIP-based XLSX file")
    return errors


def LoadOpenpyxl():
    try:
        import defusedxml  # noqa: F401
        import openpyxl
    except ImportError as error:
        raise RuntimeError(
            "install tools/etl/requirements.txt in an isolated environment"
        ) from error
    return openpyxl


def InspectTables(path: Path) -> tuple[dict[str, object], list[str]]:
    openpyxl = LoadOpenpyxl()
    from openpyxl.utils.cell import range_boundaries

    workbook = openpyxl.load_workbook(
        path,
        read_only=False,
        data_only=False,
        keep_links=False,
    )
    found: dict[str, object] = {}
    errors: list[str] = []
    for worksheet in workbook.worksheets:
        for table_name in worksheet.tables.keys():
            if table_name not in EXPECTED_TABLES:
                continue
            table = worksheet.tables[table_name]
            minimum_column, minimum_row, maximum_column, maximum_row = (
                range_boundaries(table.ref)
            )
            headers = tuple(
                worksheet.cell(minimum_row, column).value
                for column in range(minimum_column, maximum_column + 1)
            )
            if headers != EXPECTED_TABLES[table_name]:
                errors.append(f"{table_name} headers do not match the contract")
            identifiers: list[str] = []
            formula_count = 0
            for row in worksheet.iter_rows(
                min_row=minimum_row + 1,
                max_row=maximum_row,
                min_col=minimum_column,
                max_col=maximum_column,
            ):
                identifier = row[0].value
                identifiers.append("" if identifier is None else str(identifier).strip())
                formula_count += sum(
                    isinstance(cell.value, str) and cell.value.startswith("=")
                    for cell in row
                )
            blank_identifiers = sum(not identifier for identifier in identifiers)
            duplicate_identifiers = len(identifiers) - len(set(identifiers))
            if blank_identifiers:
                errors.append(f"{table_name} contains blank row identifiers")
            if duplicate_identifiers:
                errors.append(f"{table_name} contains duplicate row identifiers")
            found[table_name] = {
                "sheet": worksheet.title,
                "range": table.ref,
                "rows": len(identifiers),
                "formula_cells": formula_count,
                "blank_identifiers": blank_identifiers,
                "duplicate_identifiers": duplicate_identifiers,
            }
    missing = sorted(set(EXPECTED_TABLES) - set(found))
    if missing:
        errors.append("missing required tables: " + ", ".join(missing))
    workbook.close()
    return found, errors


def FileDigest(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as workbook_file:
        for block in iter(lambda: workbook_file.read(1024 * 1024), b""):
            digest.update(block)
    return digest.hexdigest()


def Main() -> int:
    arguments = ParseArguments()
    errors = PreflightArchive(arguments.workbook)
    tables: dict[str, object] = {}
    if not errors:
        try:
            tables, table_errors = InspectTables(arguments.workbook)
            errors.extend(table_errors)
        except RuntimeError as error:
            errors.append(str(error))
    report = {
        "valid": not errors,
        "sha256": FileDigest(arguments.workbook)
        if arguments.workbook.is_file()
        else None,
        "tables": tables,
        "errors": errors,
    }
    rendered = json.dumps(report, indent=2, sort_keys=True) + "\n"
    if arguments.report:
        arguments.report.parent.mkdir(parents=True, exist_ok=True)
        arguments.report.write_text(rendered, encoding="utf-8")
    else:
        sys.stdout.write(rendered)
    return 0 if not errors else 1


if __name__ == "__main__":
    raise SystemExit(Main())
